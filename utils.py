import os
os.environ['CUDA_VISIBLE_DEVICES'] = '0,1,2'
import json
import numpy as np
import openpyxl
import torch
import models
import torch.optim as optim
from PIL import Image
from sklearn.metrics import f1_score, roc_auc_score, precision_score, recall_score, roc_curve


def denormalize(T, coords):
    return (0.5 * ((coords + 1.0) * T))

class AverageMeter(object):
    """
    Computes and stores the average and
    current value.
    """
    def __init__(self):
        self.reset()

    def reset(self):
        self.val = 0
        self.avg = 0
        self.sum = 0
        self.count = 0

    def update(self, val, n=1):
        self.val = val
        self.sum += val * n
        self.count += n
        self.avg = self.sum / self.count

def accuracy(output, target):
    """Computes the precision@k for the specified values of k"""

    batch_size = target.size(0)
    preds = output.argmax(dim=1)
    correct_k = torch.sum(preds.view(-1) == target.view(-1)).item()

    correct_k = correct_k / batch_size

    # print("output in accuracy:======================================================================")
    # print(output)
    # print("preds in accuracy:=======================================================================")
    # print(preds)
    # print("correct_k:===============================================================================")
    # print(correct_k)

    return correct_k


def f1(output, target):
    preds = output.argmax(dim=1)
    return f1_score(target.data.cpu().numpy(), preds.data.cpu().numpy())


def precision(output, target):
    preds = output.argmax(dim=1)
    return precision_score(target.data.cpu().numpy(), preds.data.cpu().numpy())


def recall(output, target):
    preds = output.argmax(dim=1)
    return recall_score(target.data.cpu().numpy(), preds.data.cpu().numpy())


def auc(output, target):
    preds = output.argmax(dim=1)
    return roc_auc_score(target.data.cpu().numpy(), preds.data.cpu().numpy())


def roc(output, target):
    preds = output.argmax(dim=1)
    return roc_curve(target.data.cpu().numpy(), preds.data.cpu().numpy(), pos_label=1)


def f1_ronghe(output, target):
    preds = output.argmax(dim=1)
    return f1_score(target.data.cpu().numpy(), preds.data.cpu().numpy(), average='macro')


def precision_ronghe(output, target):
    preds = output.argmax(dim=1)
    return precision_score(target.data.cpu().numpy(), preds.data.cpu().numpy(), average='macro')


def recall_ronghe(output, target):
    preds = output.argmax(dim=1)
    return recall_score(target.data.cpu().numpy(), preds.data.cpu().numpy(), average='macro')


def auc_ronghe(output, target):
    preds = output.argmax(dim=1)
    return roc_auc_score(target.data.cpu().numpy(), preds.data.cpu().numpy(), average='samples')

def resize_array(x, size):
    # 3D and 4D tensors allowed only
    assert x.ndim in [3, 4], "Only 3D and 4D Tensors allowed!"

    # 4D Tensor
    if x.ndim == 4:
        res = []
        for i in range(x.shape[0]):
            img = array2img(x[i])
            img = img.resize((size, size))
            img = np.asarray(img, dtype='float32')
            img = np.expand_dims(img, axis=0)
            img /= 255.0
            res.append(img)
        res = np.concatenate(res)
        res = np.expand_dims(res, axis=1)
        return res

    # 3D Tensor
    img = array2img(x)
    img = img.resize((size, size))
    res = np.asarray(img, dtype='float32')
    res = np.expand_dims(res, axis=0)
    res /= 255.0
    return res


def img2array(data_path, desired_size=None, expand=False, view=False):
    """
    Util function for loading RGB image into a numpy array.

    Returns array of shape (1, H, W, C).
    """
    img = Image.open(data_path)
    img = img.convert('RGB')
    if desired_size:
        img = img.resize((desired_size[1], desired_size[0]))
    if view:
        img.show()
    x = np.asarray(img, dtype='float32')
    if expand:
        x = np.expand_dims(x, axis=0)
    x /= 255.0
    return x


def array2img(x):
    """
    Util function for converting anumpy array to a PIL img.

    Returns PIL RGB img.
    """
    x = np.asarray(x)
    x = x + max(-np.min(x), 0)
    x_max = np.max(x)
    if x_max != 0:
        x /= x_max
    x *= 255
    return Image.fromarray(x.astype('uint8'), 'RGB')

def prepare_dirs(config):
    for path in [config.ckpt_dir, config.logs_dir]:
        if not os.path.exists(path):
            os.makedirs(path)


def save_config(config):
    model_name = config.save_name
    filename = model_name + '_params.json'
    param_path = os.path.join(config.ckpt_dir, filename)

    print("[*] Model Checkpoint Dir: {}".format(config.ckpt_dir))
    print("[*] Param Path: {}".format(param_path))

    with open(param_path, 'w') as fp:
        json.dump(config.__dict__, fp, indent=4, sort_keys=True)


def loader_model(num_classes, name, path):
    checkpoint = torch.load(path)
    if num_classes==2:
        if name=='googlenet':
            model = models.get_googlenet(2, False, False, 1)
        elif name=='resnet50':
            model = models.get_resnet50(2, False, False, 1)
        elif name=='resnet34':
            model = models.get_resnet34(2, False, False, 1)
        elif name == 'se_resnet50':
            model = models.get_se_resnet50(2, False, False, 1)
        elif name == 'se_resnext101':
            model = models.get_se_resnetxt101(2, False, False, 1)
    elif num_classes == 9:
        if name == 'googlenet':
            model = models.get_googlenet(8, False, False, 1)
        elif name == 'resnet50':
            model = models.get_resnet50(8, False, False, 1)
        elif name == 'resnet34':
            model = models.get_resnet34(8, False, False, 1)
        elif name == 'se_resnet50':
            model = models.get_se_resnet50(9, False, False, 1)
        elif name == 'se_resnext101':
            model = models.get_se_resnetxt101(9, False, False, 1)
        elif name == 'se_resnet152':
            model = models.get_se_resnet152(9, False, False, 1)
        elif name == 'senet154':
            model = models.get_senet154(9, False, False, 1)
        elif name == 'efficientnetb5':
            model = models.get_efficientnetb5(9, False, False, 1)
        elif name == 'efficientnetb6':
            model = models.get_efficientnetb6(9, False, False, 1)
        elif name == 'efficientnetb7':
            model = models.get_efficientnetb7(9, False, False, 1)
        elif name == 'adv_efficientnetb5':
            model = models.get_advefficientnetb5(9, False, False, 1)
        elif name == 'adv_efficientnetb6':
            model = models.get_advefficientnetb6(9, False, False, 1)
        elif name == 'adv_efficientnetb7':
            model = models.get_advefficientnetb7(9, False, False, 1)
        elif name == 'adv_efficientnetb8':
            model = models.get_advefficientnetb8(9, False, False, 1)
        elif name == 'wsl_328':
            model = models.get_resnext101wsl328(9, False, False, 1)
        elif name == 'wsl_3216':
            model = models.get_resnext101wsl3216(9, False, False, 1)
        elif name == 'se_resnext50':
            model = models.get_se_resnetxt50(9, False, False, 1)
        elif name == 'se_resnet101':
            model = models.get_se_resnet101(9, False, False, 1)
    elif num_classes == 6:
        if name == 'googlenet':
            model = models.get_googlenet(6, False, False, 1)
        elif name == 'resnet50':
            model = models.get_resnet50(6, False, False, 1)
        elif name == 'resnet34':
            model = models.get_resnet34(6, False, False, 1)
        elif name == 'se_resnet50':
            model = models.get_se_resnet50(6, False, False, 1)
        elif name == 'se_resnext101':
            model = models.get_se_resnetxt101(6, False, False, 1)
        elif name == 'se_resnet152':
            model = models.get_se_resnet152(6, False, False, 1)
        elif name == 'senet154':
            model = models.get_senet154(6, False, False, 1)
        elif name == 'efficientnetb5':
            model = models.get_efficientnetb5(6, False, False, 1)
        elif name == 'efficientnetb6':
            model = models.get_efficientnetb6(6, False, False, 1)
        elif name == 'efficientnetb7':
            model = models.get_efficientnetb7(6, False, False, 1)
        elif name == 'adv_efficientnetb5':
            model = models.get_advefficientnetb5(6, False, False, 1)
        elif name == 'adv_efficientnetb6':
            model = models.get_advefficientnetb6(6, False, False, 1)
        elif name == 'adv_efficientnetb7':
            model = models.get_advefficientnetb7(6, False, False, 1)
        elif name == 'adv_efficientnetb8':
            model = models.get_advefficientnetb8(6, False, False, 1)
        elif name == 'wsl_328':
            model = models.get_resnext101wsl328(6, False, False, 1)
        elif name == 'wsl_3216':
            model = models.get_resnext101wsl3216(6, False, False, 1)
        elif name == 'se_resnext50':
            model = models.get_se_resnetxt50(6, False, False, 1)
        elif name == 'se_resnet101':
            model = models.get_se_resnet101(6, False, False, 1)
    # checkpoint_dict = checkpoint.module.state_dict()
    # model.load_state_dict(checkpoint['model_state'], strict=False)
    # if torch.cuda.device_count() > 1:
    #     model = torch.nn.DataParallel(model, device_ids=[0])
    # model = model.cuda()
    # model.load_state_dict(checkpoint_dict)
    # model = torch.nn.DataParallel(model)
    # model = torch.nn.DataParallel(model)
    # model = torch.nn.DataParallel(model).module()
    model.load_state_dict(checkpoint['model_state'])
    # model.load_state_dict({k.replace('module.', ''): v for k, v in checkpoint.items()})
    # model = model.cuda()
    # model.eval().cuda()
    optimizer = optim.SGD(model.parameters(), lr=0.01)
    optimizer.load_state_dict(checkpoint['optim_state'])
    epoch = checkpoint['epoch']
    return model, optimizer, epoch


def getWorkBook():
    mywb = openpyxl.Workbook()
    mywb.create_sheet(index=0,title='epoch_trainloss')
    mywb.create_sheet(index=1,title='epoch_trainacc')
    mywb.create_sheet(index=2, title='epoch_trainf1')
    mywb.create_sheet(index=3, title='epoch_trainprecision')
    mywb.create_sheet(index=4, title='epoch_trainrecall')
    mywb.create_sheet(index=5, title='epoch_trainauc')
    mywb.create_sheet(index=6,title='epoch_testloss')
    mywb.create_sheet(index=7,title='epoch_testacc')
    mywb.create_sheet(index=8, title='epoch_testf1')
    mywb.create_sheet(index=9, title='epoch_testprecision')
    mywb.create_sheet(index=10, title='epoch_testrecall')
    mywb.create_sheet(index=11, title='epoch_testauc')
    return mywb

def getWorkBook_DML():
    mywb = openpyxl.Workbook()
    mywb.create_sheet(index=0,title='model1_trainloss')
    mywb.create_sheet(index=1,title='model1_trainacc')
    mywb.create_sheet(index=2, title='model1_trainf1')
    mywb.create_sheet(index=3, title='model1_trainprecision')
    mywb.create_sheet(index=4, title='model1_trainrecall')
    mywb.create_sheet(index=5, title='model1_trainauc')
    mywb.create_sheet(index=6,title='model1_testloss')
    mywb.create_sheet(index=7,title='model1_testacc')
    mywb.create_sheet(index=8, title='model1_testf1')
    mywb.create_sheet(index=9, title='model1_testprecision')
    mywb.create_sheet(index=10, title='model1_testrecall')
    mywb.create_sheet(index=11, title='model1_testauc')


    mywb.create_sheet(index=12,title='model2_trainloss')
    mywb.create_sheet(index=13,title='model2_trainacc')
    mywb.create_sheet(index=14, title='model2_trainf1')
    mywb.create_sheet(index=15, title='model2_trainprecision')
    mywb.create_sheet(index=16, title='model2_trainrecall')
    mywb.create_sheet(index=17, title='model2_trainauc')
    mywb.create_sheet(index=18,title='model2_testloss')
    mywb.create_sheet(index=19,title='model2_testacc')
    mywb.create_sheet(index=20, title='model2_testf1')
    mywb.create_sheet(index=21, title='model2_testprecision')
    mywb.create_sheet(index=22, title='model2_testrecall')
    mywb.create_sheet(index=23, title='model2_testauc')

    return mywb
def loader_model1(num_classes, name, path):

    # global model
    checkpoint = torch.load(path)
    if num_classes==2:
        if name=='googlenet':
            model = models.get_googlenet(2, False, False, 1)
        elif name=='resnet50':
            model = models.get_resnet50(2, False, False, 1)
        elif name=='resnet34':
            model = models.get_resnet34(2, False, False, 1)
        elif name == 'se_resnet50':
            model = models.get_se_resnet50(2, False, False, 1)
            model.load_state_dict(checkpoint['model_state'])

        elif name == 'se_resnext101':
            model = models.get_se_resnetxt101(2, False, False, 1)
            model.load_state_dict(checkpoint['model_state'])
    elif num_classes==9:
        if name=='googlenet':
            model = models.get_googlenet(9, False, False, 1)
        elif name=='resnet50':
            model = models.get_resnet50(9, False, False, 1)
        elif name=='resnet34':
            model = models.get_resnet34(9, False, False, 1)
        elif name == 'se_resnet50':
            model = models.get_se_resnet50(9, False, False, 1)
            model.load_state_dict(checkpoint['model_state'])

        elif name == 'se_resnext101':
            model = models.get_se_resnetxt101(9, False, False, 1)
            model.load_state_dict(checkpoint['model_state'])
        # elif name == 'se_resnet152':
        #     model = models.get_se_resnet152(9, False, False, 1)
        #     model.load_state_dict(checkpoint['model_state'])
        # elif name =='senet154':
        #     model = models.get_senet154(9, False, False, 1)
        # elif name == 'efficientnetb5':
        #     model = models.get_efficientnetb5(9, False, False, 1)
        # elif name == 'efficientnetb6':
        #     model = models.get_efficientnetb6(9, False, False, 1)
        # elif name == 'efficientnetb7':
        #     model = models.get_efficientnetb7(9, False, False, 1)
        # elif name == 'adv_efficientnetb5':
        #     model = models.get_advefficientnetb5(9, False, False, 1)
        # elif name == 'adv_efficientnetb6':
        #     model = models.get_advefficientnetb6(9, False, False, 1)
        # elif name == 'adv_efficientnetb7':
        #     model = models.get_advefficientnetb7(9, False, False, 1)
        # elif name == 'adv_efficientnetb8':
        #     model = models.get_advefficientnetb8(9, False, False, 1)
        # elif name == 'wsl_328':
        #     model = models.get_resnext101wsl328(9, False, False, 1)
        # elif name == 'wsl_3216':
        #     model = models.get_resnext101wsl3216(9, False, False, 1)
        # elif name == 'se_resnext50':
        #     model = models.get_se_resnetxt50(9, False, False, 1)
        #     model.load_state_dict(checkpoint['model_state'])
        # elif name == 'se_resnet101':
        #     model = models.get_se_resnet101(9, False, False, 1)
        #     model.load_state_dict(checkpoint['model_state'])
    elif num_classes==1000:
        if name=='googlenet':
            model = models.get_googlenet(1000, True, False, 1)
        elif name=='resnet50':
            model = models.get_resnet50(1000, True, False, 1)
        elif name=='resnet34':
            model = models.get_resnet34(1000, True, False, 1)
        elif name == 'se_resnet50':
            model = models.get_se_resnet50(1000, True, False, 1)
        elif name == 'se_resnext101':
            model = models.get_se_resnetxt101(1000, True, False, 1)
        # elif name == 'se_resnet152':
        #     model = models.get_se_resnet152(1000, True, False, 1)
        # elif name =='senet154':
        #     model = models.get_senet154(1000, True, False, 1)
        # elif name == 'efficientnetb5':
        #     model = models.get_efficientnetb5(1000, True, False, 1)
        # elif name == 'efficientnetb6':
        #     model = models.get_efficientnetb6(1000, True, False, 1)
        # elif name == 'efficientnetb7':
        #     model = models.get_efficientnetb7(1000, True, False, 1)
        # elif name == 'adv_efficientnetb5':
        #     model = models.get_advefficientnetb5(1000, True, False, 1)
        # elif name == 'adv_efficientnetb6':
        #     model = models.get_advefficientnetb6(1000, True, False, 1)
        # elif name == 'adv_efficientnetb7':
        #     model = models.get_advefficientnetb7(1000, True, False, 1)
        # elif name == 'adv_efficientnetb8':
        #     model = models.get_advefficientnetb8(1000, True, False, 1)
        # elif name == 'wsl_328':
        #     model = models.get_resnext101wsl328(1000, True, False, 1)
        # elif name == 'wsl_3216':
        #     model = models.get_resnext101wsl3216(1000, True, False, 1)
        # elif name == 'se_resnext50':
        #     model = models.get_se_resnetxt50(1000, True, False, 1)
        # elif name == 'se_resnet101':
        #     model = models.get_se_resnet101(1000, True, False, 1)
    elif num_classes == 6:
        if name == 'googlenet':
            model = models.get_googlenet(6, False, False, 1)
        elif name == 'resnet50':
            model = models.get_resnet50(6, False, False, 1)
        elif name == 'resnet34':
            model = models.get_resnet34(6, False, False, 1)
        elif name == 'se_resnet50':
            model = models.get_se_resnet50(6, False, False, 1)
        elif name == 'se_resnext101':
            model = models.get_se_resnetxt101(6, False, False, 1)
        # elif name == 'se_resnet152':
        #     model = models.get_se_resnet152(6, False, False, 1)
        # elif name =='senet154':
        #     model = models.get_senet154(6, False, False, 1)
        # elif name == 'efficientnetb5':
        #     model = models.get_efficientnetb5(6, False, False, 1)
        # elif name == 'efficientnetb6':
        #     model = models.get_efficientnetb6(6, False, False, 1)
        # elif name == 'efficientnetb7':
        #     model = models.get_efficientnetb7(6, False, False, 1)
        # elif name == 'adv_efficientnetb5':
        #     model = models.get_advefficientnetb5(6, False, False, 1)
        # elif name == 'adv_efficientnetb6':
        #     model = models.get_advefficientnetb6(6, False, False, 1)
        # elif name == 'adv_efficientnetb7':
        #     model = models.get_advefficientnetb7(6, False, False, 1)
        # elif name == 'adv_efficientnetb8':
        #     model = models.get_advefficientnetb8(6, False, False, 1)
        # elif name == 'wsl_328':
        #     model = models.get_resnext101wsl328(6, False, False, 1)
        # elif name == 'wsl_3216':
        #     model = models.get_resnext101wsl3216(6, False, False, 1)
        # elif name == 'se_resnext50':
        #     model = models.get_se_resnetxt50(6, False, False, 1)
        # elif name == 'se_resnet101':
        #     model = models.get_se_resnet101(6, False, False, 1)
    # checkpoint_dict = checkpoint.module.state_dict()
    # model.load_state_dict(checkpoint['model_state'], strict=False)
    # if torch.cuda.device_count() > 1:
    #     model = torch.nn.DataParallel(model, device_ids=[0])
    # model = model.cuda()
    # model.load_state_dict(checkpoint_dict)
    # model = torch.nn.DataParallel(model)
    # model = torch.nn.DataParallel(model)
    # model = torch.nn.DataParallel(model).module()
    model.load_state_dict(checkpoint['model_state'])
    # model.load_state_dict(checkpoint)
    # model.load_state_dict({k.replace('module.', ''): v for k, v in checkpoint.items()})
    # model = model.cuda()
    # model.eval().cuda()
    optimizer = optim.SGD(model.parameters(), lr=0.01)
    optimizer.load_state_dict(checkpoint['optim_state'])
    epoch = checkpoint['epoch']
    return model,optimizer,epoch